from django.shortcuts import render
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt, csrf_protect
from django.views.decorators.http import require_POST
import json
from openpyxl import Workbook, load_workbook
import os
import re
from difflib import get_close_matches

KEY_PAIR_RESPONSES = {
    "hello": "Hello! How can I help you today?",
    "how are you": "I'm just a bot, but I'm doing great, thank you!",
    "what can you do": "I'm here to help you with any questions you have. Just ask me anything!",
    "who are you": "I'm a bot created by Mohammed Shaik Sahil. Nice to meet you!",
    "bye": "Goodbye! Have a great day!",
    "what is your name": "I'm a bot, you can call me Jack.",
    "help": "I'm here to help you! Just ask me anything."
}
ADMIN_COMMANDS = ["/admin", "/show", "/clear", "/end", "/register", "/stop", "/cmd", "/cmd-admin"]

SECRET_KEY = "admin123"

user_admin_state = {}
user_registration_state = {}

from openpyxl import Workbook, load_workbook
import os

def save_to_excel(name, mobile, email, course, class_name, roll_number, event_name='Default Event'):
    file_path = "register.xlsx"
    
    # Check if the file exists
    if not os.path.exists(file_path):
        wb = Workbook()
        ws = wb.active
        ws.title = event_name
        ws.append(['Name', 'Mobile', 'Email', 'Course', 'Class Name', 'Roll Number'])
    else:
        wb = load_workbook(file_path)
        
        # Check if the event sheet exists, if not create it
        if event_name in wb.sheetnames:
            ws = wb[event_name]
        else:
            # If the event name doesn't exist, use the first sheet or add a new one
            ws = wb.active  # Use the first sheet
            # Only append headers if the sheet is empty (i.e., no rows except the header)
            if not any(ws.iter_rows(min_row=1, max_row=1, values_only=True)):  # If no rows exist, append headers
                ws.append(['Name', 'Mobile', 'Email', 'Course', 'Class Name', 'Roll Number'])

        # Check for duplicate registration
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] == name and str(row[1]) == mobile:
                return "already_registered"

    # Append the new registration data to the sheet
    ws.append([name, mobile, email, course, class_name, roll_number])
    wb.save(file_path)
    return "registered"

def read_from_excel():
    file_path = "register.xlsx"
    
    if not os.path.exists(file_path):
        return "No registrations found."

    wb = load_workbook(file_path)
    ws = wb.active
    data = []

    for row in ws.iter_rows(values_only=True):
        data.append(row)

    return data

def find_closest_word(input_word, word_list):
    closest_matches = get_close_matches(input_word, word_list)
    return closest_matches[0] if closest_matches else None

@csrf_exempt
@require_POST
def get_bot_response(request):
    try:
        data = json.loads(request.body)
        user_message = data.get('message').strip().lower()
        user_id = data.get('user_id')

        if user_id not in user_admin_state:
            user_admin_state[user_id] = {
                'is_admin': False,
                'attempts': 0
            }
        if user_id not in user_registration_state:
            user_registration_state[user_id] = {
                'registration_step': 0,
                'name': '',
                'mobile': '',
                'email': '',
                'course': '',
                'class_name': '',
                'roll_number': ''
            }

        # Handle /register command
        if user_message == "/register":
            user_registration_state[user_id]['registration_step'] = 1
            return JsonResponse({'response': "Please enter your name:"})

        # Handle /stop command to stop the registration process
        if user_message == "/stop":
            if user_registration_state[user_id]['registration_step'] != 0:
                user_registration_state[user_id]['registration_step'] = 0
                user_registration_state[user_id] = {
                    'registration_step': 0,
                    'name': '',
                    'mobile': '',
                    'email': '',
                    'course': '',
                    'class_name': '',
                    'roll_number': ''
                }
                return JsonResponse({'response': "Registration has been stopped and all entered data has been cleared."})
            else:
                return JsonResponse({'response': "No registration is currently in progress."})

        # Handle /cmd to show user-related commands
        if user_message == "/cmd":
            commands = {
                "/register": "To start the registration process and register user details.",
                "/stop": "To stop the current registration process and clear entered data.",
                "/cmd": "To show the list of user-related commands.",
                "/clear": "To clear the chat or reset the session."
            }
            command_list = "\n".join([f"{cmd}: {desc}" for cmd, desc in commands.items()])
            return JsonResponse({'response': f"User-related commands:\n{command_list}"})

        # Handle /cmd-admin to show admin-related commands
        if user_message == "/cmd-admin":
            admin_commands = {
                "/admin": "To enter admin mode with the correct secret key.",
                "/show": "To view the data in the registration table.",
                "/end": "To exit admin mode.",
                "/clear": "To clear the chat or reset the session.",
            }
            admin_command_list = "\n".join([f"{cmd}: {desc}" for cmd, desc in admin_commands.items()])
            return JsonResponse({'response': f"Admin-related commands:\n{admin_command_list}"})

        if user_message.startswith("/"):
            # Check for closest match for admin commands, but only if the user is an admin
            if user_admin_state[user_id]['is_admin']:
                closest_command = find_closest_word(user_message, ADMIN_COMMANDS)
                if closest_command and closest_command != user_message:
                    return JsonResponse({'response': f"Did you mean '{closest_command}'?"})

            if user_message.startswith("/admin"):
                if user_admin_state[user_id]['is_admin']:
                    return JsonResponse({'response': "You are already in admin mode."})
                user_admin_state[user_id]['is_admin'] = True
                user_admin_state[user_id]['attempts'] = 0
                return JsonResponse({'response': "Please enter the secret key to enter admin mode."})
            
            elif user_message.startswith("/end") and user_admin_state[user_id]['is_admin']:
                user_admin_state[user_id]['is_admin'] = False
                return JsonResponse({'response': "You have exited admin mode."})

            elif user_message.startswith("/show") and user_admin_state[user_id]['is_admin'] == "granted":
                data = read_from_excel()
                if data == "No registrations found.":
                    return JsonResponse({'response': data})
                # Convert data to HTML table
                table_html = "<table border='1'>"
                for row in data:
                    table_html += "<tr>"
                    for cell in row:
                        table_html += f"<td>{cell}</td>"
                    table_html += "</tr>"
                table_html += "</table>"
                return JsonResponse({'response': table_html})
            
            elif user_message.startswith("/clear") and user_admin_state[user_id]['is_admin'] == "granted":
                return JsonResponse({'response': 'Cleared'})

        # If in admin mode but not yet confirmed
        if user_admin_state[user_id]['is_admin'] and user_admin_state[user_id]['is_admin'] != "granted":
            if user_message == SECRET_KEY:
                user_admin_state[user_id]['is_admin'] = "granted"
                return JsonResponse({'response': "Hey admin, nice to meet you!"})
            else:
                user_admin_state[user_id]['attempts'] += 1
                # If user fails 3 times, exit admin mode
                if user_admin_state[user_id]['attempts'] >= 3:
                    user_admin_state[user_id]['is_admin'] = False
                    return JsonResponse({'response': "Admin mode failed. You have been logged out."})
                return JsonResponse({'response': f"Incorrect secret key. You have {3 - user_admin_state[user_id]['attempts']} attempts left."})
        
        # Normal bot functionality (non-admin mode)

        if user_admin_state[user_id]['is_admin'] != "granted":
            registration_step = user_registration_state[user_id]['registration_step']

            if registration_step == 1:
                user_message = user_message.strip().upper()  # Convert to uppercase and strip whitespace
                if not user_message:
                    return JsonResponse({'response': "Name cannot be empty. Please enter a valid name."})
                
                user_registration_state[user_id]['name'] = user_message
                user_registration_state[user_id]['registration_step'] = 2
                return JsonResponse({'response': "Please enter your mobile number:"})

            if registration_step == 2:
                if not user_message.isdigit() or len(user_message) != 10:
                    return JsonResponse({'response': "Please enter a valid 10-digit mobile number without any other characters."})

                user_registration_state[user_id]['mobile'] = user_message
                user_registration_state[user_id]['registration_step'] = 3
                return JsonResponse({'response': "Please enter your email:"})

            if registration_step == 3:
                user_message = user_message.strip().lower()  # Convert input to lowercase and remove leading/trailing spaces

                if not user_message.endswith("@gmail.com"):
                    return JsonResponse({'response': "Please enter a valid email address ending with @gmail.com."})

                user_registration_state[user_id]['email'] = user_message
                user_registration_state[user_id]['registration_step'] = 4
                return JsonResponse({'response': "Please enter your course name, BE or DIPLOMA:"})

            if registration_step == 4:
                user_message = user_message.strip().upper()  # Convert input to uppercase and remove any leading/trailing spaces

                if user_message not in ["BE", "DIPLOMA"] and len(user_message.split()) > 2:
                    return JsonResponse({'response': "Please enter a valid course name: BE or DIPLOMA, in all caps, with max 1 space."})

                user_registration_state[user_id]['course'] = user_message
                user_registration_state[user_id]['registration_step'] = 5
                return JsonResponse({'response': "Please enter your class name (1 space only, no special characters):"})

            if registration_step == 5:
                user_message = user_message.strip().upper()  # Convert input to uppercase and remove leading/trailing spaces
                
                if not re.match(r'^[A-Z]+\s?[A-Z]+$', user_message):
                    return JsonResponse({'response': "Please enter a valid class name: 1 space only, no special characters, and all in uppercase."})

                user_registration_state[user_id]['class_name'] = user_message
                user_registration_state[user_id]['registration_step'] = 6
                return JsonResponse({'response': "Please enter your roll number (starting with 1610):"})

            if registration_step == 6:
                if not user_message.startswith("1610"):
                    return JsonResponse({'response': "Please enter a valid roll number starting with 1610."})

                user_registration_state[user_id]['roll_number'] = user_message
                user_registration_state[user_id]['registration_step'] = 0
                user_name = user_registration_state[user_id]['name']
                user_mobile = user_registration_state[user_id]['mobile']
                user_email = user_registration_state[user_id]['email']
                user_course = user_registration_state[user_id]['course']
                user_class_name = user_registration_state[user_id]['class_name']
                user_roll_number = user_registration_state[user_id]['roll_number']
                registration_status = save_to_excel(user_name, user_mobile, user_email, user_course, user_class_name, user_roll_number)

                if registration_status == "already_registered":
                    return JsonResponse({'response': "You are already registered with this name and mobile number."})
                else:
                    if user_id in user_registration_state:
                        del user_registration_state[user_id]
                    try:
                        wb = load_workbook("register.xlsx")
                        ws = wb.active
                        event_name = ws.title
                    except:
                        event_name = "Default Event"
                    return JsonResponse({'response': f"User registered for event {event_name} with details: name: {user_name}, mobile: {user_mobile}, email: {user_email}, course: {user_course}, class name: {user_class_name}, roll number: {user_roll_number}."})

        else:
            return JsonResponse({'response': "You are an admin and cannot perform user registration."})

        closest_word = find_closest_word(user_message, KEY_PAIR_RESPONSES.keys())
        if closest_word and closest_word != user_message:
            return JsonResponse({'response': f"Did you mean '{closest_word}'?"})
        
        bot_response = KEY_PAIR_RESPONSES.get(user_message, "Sorry, I couldn't understand that. Could you please rephrase?")
        return JsonResponse({'response': bot_response})

    except Exception as e:
        return JsonResponse({'error': 'An error occurred while processing your request. Please try again later.'}, status=500)
    
def index(request):
    return render(request, 'index.html')

def message(request):
    return render(request, 'message.html')
